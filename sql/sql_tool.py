from distutils.log import debug
import pymysql
import yaml
import sys
import argparse
import openpyxl
import pandas
from loguru import logger
import time

def banner():
    print(r"""
 ________        ________        ___            _________     ________       ________       ___           
|\   ____\      |\   __  \      |\  \          |\___   ___\  |\   __  \     |\   __  \     |\  \          
\ \  \___|_     \ \  \|\  \     \ \  \         \|___ \  \_|  \ \  \|\  \    \ \  \|\  \    \ \  \         
 \ \_____  \     \ \  \\\  \     \ \  \             \ \  \    \ \  \\\  \    \ \  \\\  \    \ \  \        
  \|____|\  \     \ \  \\\  \     \ \  \____         \ \  \    \ \  \\\  \    \ \  \\\  \    \ \  \____   
    ____\_\  \     \ \_____  \     \ \_______\        \ \__\    \ \_______\    \ \_______\    \ \_______\ 
   |\_________\     \|___| \__\     \|_______|         \|__|     \|_______|     \|_______|     \|_______| 
   \|_________|           \|__|                                                                           
                                                                                                by:XIN
    """)

def load_config():
    with open('d:\pycode-learning\sql\config.yaml', 'r') as f:
        config = yaml.load(f,Loader=yaml.FullLoader)
        host = config['all_config']['host']
        port = config['all_config']['port']
        user = config['all_config']['user']
        password = config['all_config']['password']
        database = config['all_config']['database']
    return host, port, user, password, database

def parser_error(errmsg):
    print("Usage: python " + sys.argv[0] + " [Options] use -h for help")
    print("Error: " + errmsg)
    sys.exit()


def parse_args():
    parser = argparse.ArgumentParser(description='这是一个数据库查询程序')
    parser.error = parser_error
    parser._optionals.title = "OPTIONS"
    parser.add_argument('-d', help="删除内容", default="")
    parser.add_argument('-s', help="查询语句", default="")
    parser.add_argument('-i', help="excel导入数据到数据库", default="")
    parser.add_argument('-tail', help="在控制台打印查询后10行结果。", action='store_true')
    parser.add_argument('-p', help="在控制台打印查询结果,默认查询所有。", action='store_true')
    parser.add_argument('-o', help="输出查询结果至excel,默认查询所有。", action='store_true')
    return parser.parse_args()

def out_con(sql):
    # 连接数据库
    conn = pymysql.connect(host=load_config()[0], port=load_config()[1], user=load_config()[2], password=load_config()[3], charset='utf8', database=load_config()[4])
    # 使用 cursor() 方法创建一个游标对象 cursor
    cursor = conn.cursor()
    if sql =="":
        sql='select * from number'
    logger.info(sql)
    try:
        #执行sql
        cursor.execute(sql)
        #获取查询结果
        results=cursor.fetchall()
        for row in results:
            id=row[0]
            number=row[1]
            type=row[2]
            name=row[3]
            #输出
            print('id:',id,'number:',number,'type:',type,'name:',name)
    except Exception as e:
        print(e)
        logger.info('查询失败：{}'.format(e))
    finally:
        conn.close()

# 导出至excel表
def out_excel(sql):
    conn = pymysql.connect(host=load_config()[0], port=load_config()[1], user=load_config()[2], password=load_config()[3], charset='utf8', database=load_config()[4])
    if sql =="":
        sql='select * from number'
    logger.info(sql)
    try:
        with conn.cursor() as cursor:#创建游标，在这里conn.cursor()==cursor
            cursor.execute(sql)
            excel_workbook=openpyxl.Workbook()#创建一个excel工作簿
            excel_sheet=excel_workbook.active #使用默认的sheet表，不新建表,切换当前把数据写入此表
            #注意数据量比较少的时候用fetchall()，一行一行的读取，这里数据量少就直接fetchall()了
            # openpyxl操作excel是，行和列的索引从1开始，所以要+1， col_id, col_name表示索引和索引内容
            for col_id, col_name in enumerate(['序号','通知编号','事件名称','研判人']):#列索引，读取那些列的数据
                excel_sheet.cell(1,col_id+1,col_name)
            # 写入数据
            for row_id,row_emp in enumerate(cursor.fetchall()):#获得每一行的数据
                for col_id,col_value in enumerate(row_emp):#把每一行的每一列写入，加2是因为原先表头+1的基础上，再加上表头这一行，所以要+1+1=+2
                    excel_sheet.cell(row_id+2,col_id+1,col_value)
        excel_workbook.save("华能事件通知编号记录表.xlsx")
    except pymysql.MySQLError as err:#捕获异常
        logger.info('导出失败：{}'.format(err))
    finally:
        conn.close()#无论如何都要关闭连接，节省资源占用

def input_excel(filename):
    # 读取excel数据
    data = pandas.read_excel(filename, index_col='序号')
    excel_data = []
    i = 0
    for index, row in data.iterrows():
        excel_data.append([])
        excel_data[i].append(index)
        excel_data[i].append(row["通知编号"])
        excel_data[i].append(row["事件名称"])
        excel_data[i].append(row["研判人"])
        i = i+1
    conn = pymysql.connect(host=load_config()[0], port=load_config()[1], user=load_config()[2], password=load_config()[3], charset='utf8', database=load_config()[4])
    cursor = conn.cursor()
    try:
        with conn.cursor() as cursor:
            #cursor.executemany表示批量插入数据，批处理
            cursor.executemany(
                'insert into number'
                '(id,number,type,name)'
                'values'
                '(%s,%s,%s,%s)',
            excel_data
            )
        conn.commit()
        print("导入成功！")
    except pymysql.MySQLError as err:  # 捕获异常
        print(err)  # 如果出现异常，打印错误信息
        print("导入失败！")
    finally:
        conn.close()  # 无论如何都要关闭连接，节省资源占用

def del_tab(id):
    conn = pymysql.connect(host=load_config()[0], port=load_config()[1], user=load_config()[2], password=load_config()[3], charset='utf8', database=load_config()[4])
    sql='delete from number where id = {}'.format(id)
    cursor=conn.cursor()
    logger.info(sql)
    try:
        #执行sql语句
        cursor.execute(sql)
        #提交事务
        conn.commit()
        print('删除操作成功')
    except pymysql.MySQLError as err:
        conn.rollback()
        logger.info('删除失败：{}'.format(err))
    finally:
        #关闭连接
        conn.close()

def out_ten():
    conn = pymysql.connect(host=load_config()[0], port=load_config()[1], user=load_config()[2], password=load_config()[3], charset='utf8', database=load_config()[4])
    # 使用 cursor() 方法创建一个游标对象 cursor
    cursor = conn.cursor()
    sql = "select * from number order by id desc limit 10"
    logger.info(sql)
    try:
        #执行sql
        cursor.execute(sql)
        #获取查询结果
        results=cursor.fetchall()
        for row in results:
            id=row[0]
            number=row[1]
            type=row[2]
            name=row[3]
            #输出
            print('id:',id,'number:',number,'type:',type,'name:',name)
    except Exception as e:
        print(e)
        logger.info('查询失败：{}'.format(e))
    finally:
        conn.close()

if __name__ == '__main__':
    banner()
    logger.add("logger.log",format="{time} {level} {message}", rotation="3 MB", level="DEBUG")
    args = parse_args()
    if args.s :
        sql = args.s
        if args.p:
            out_con(sql)
        elif args.o:
            out_excel(sql)
    elif args.d:
        del_tab(args.d)
    elif args.p:
        sql = ""
        out_con(sql)
    elif args.o:
        sql = ""
        out_excel(sql)
    elif args.i:
        input_excel(args.i)
    elif args.tail:
        out_ten()
    else:
        print("输入的参数有误，请使用-h参数查看帮助信息！")